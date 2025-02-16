import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import openpyxl
from openpyxl.styles import Font, PatternFill
import ttkbootstrap as tb

class ExcelApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Kaluwal Excel Viewer & Editor")
        self.root.geometry("1300x810")

        self.style = tb.Style("darkly")

        self.root.grid_columnconfigure(1, weight=1)
        self.root.grid_rowconfigure(0, weight=1)

        self.sidebar = ttk.Frame(self.root, padding=10)
        self.sidebar.grid(row=0, column=0, sticky="nsw", padx=10, pady=10)

        self.fields = ["Name", "Age", "Phone", "Email", "Address", "City", "Start Date", "End Date"]
        self.entries = {}

        for field in self.fields:
            ttk.Label(self.sidebar, text=f"{field}:").grid(sticky="w", padx=5, pady=2)
            entry = ttk.Entry(self.sidebar)
            entry.grid(sticky="ew", padx=5, pady=2)
            self.entries[field] = entry

        ttk.Label(self.sidebar, text="Subscription:").grid(sticky="w", padx=5, pady=2)
        self.subscription_var = tk.StringVar(value="Subscribed")
        self.subscription_menu = ttk.Combobox(self.sidebar, textvariable=self.subscription_var,
                                              values=["Subscribed", "Unsubscribed"])
        self.subscription_menu.grid(sticky="ew", padx=5, pady=2)

        self.employment_var = tk.BooleanVar()
        self.employment_check = ttk.Checkbutton(self.sidebar, text="Employed", variable=self.employment_var)
        self.employment_check.grid(sticky="w", padx=5, pady=2)

        button_frame = ttk.Frame(self.sidebar)
        button_frame.grid(sticky="ew", padx=5, pady=5)
        button_frame.grid_columnconfigure(0, weight=1)

        self.insert_button = ttk.Button(button_frame, text="Insert", command=self.insert_data)
        self.insert_button.grid(sticky="ew", padx=5, pady=2)

        self.edit_button = ttk.Button(button_frame, text="Edit Selected", command=self.edit_data)
        self.edit_button.grid(sticky="ew", padx=5, pady=2)

        self.delete_button = ttk.Button(button_frame, text="Delete Selected", command=self.delete_data)
        self.delete_button.grid(sticky="ew", padx=5, pady=2)

        self.load_button = ttk.Button(button_frame, text="Load Excel", command=self.load_excel)
        self.load_button.grid(sticky="ew", padx=5, pady=2)

        self.table_frame = ttk.Frame(self.root)
        self.table_frame.grid(row=0, column=1, sticky="nsew", padx=10, pady=10)

        self.columns = self.fields + ["Subscription", "Employment"]
        self.tree = ttk.Treeview(self.table_frame, columns=self.columns, show='headings')

        for col in self.columns:
            self.tree.column(col, anchor="center", width=120, stretch=True)
            self.tree.heading(col, text=col, anchor="center")

        self.tree.grid(row=0, column=0, sticky="nsew")

        y_scrollbar = ttk.Scrollbar(self.table_frame, orient="vertical", command=self.tree.yview)
        y_scrollbar.grid(row=0, column=1, sticky="ns")
        self.tree.configure(yscroll=y_scrollbar.set)

        x_scrollbar = ttk.Scrollbar(self.table_frame, orient="horizontal", command=self.tree.xview)
        x_scrollbar.grid(row=1, column=0, sticky="ew")
        self.tree.configure(xscroll=x_scrollbar.set)

        self.table_frame.grid_columnconfigure(0, weight=1)
        self.table_frame.grid_rowconfigure(0, weight=1)

        self.tree.bind("<ButtonRelease-1>", self.select_item)

        self.filepath = None

    def load_excel(self):
        filepath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if not filepath:
            return

        self.filepath = filepath
        self.tree.delete(*self.tree.get_children())

        try:
            wb = openpyxl.load_workbook(filepath)
            sheet = wb.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                self.tree.insert("", tk.END, values=row)

            wb.close()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load file: {e}")

    def insert_data(self):
        if not self.filepath:
            messagebox.showwarning("Warning", "Please load an Excel file first.")
            return

        values = [self.entries[field].get() for field in self.entries]
        values.append(self.subscription_var.get())
        values.append("Employed" if self.employment_var.get() else "Unemployed")

        if not values[0] or not values[1].isdigit():
            messagebox.showerror("Input Error", "Please enter a valid Name and Age.")
            return

        try:
            wb = openpyxl.load_workbook(self.filepath)
            sheet = wb.active

            if sheet.max_row == 1:
                sheet.append(self.columns)  # Add headers if not present
                self.format_headers(sheet)

            sheet.append(values)
            wb.save(self.filepath)
            wb.close()

            self.tree.insert("", tk.END, values=values)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to insert data: {e}")

    def select_item(self, event):
        selected = self.tree.selection()
        if not selected:
            return

        item = self.tree.item(selected[0], "values")

        for i, field in enumerate(self.entries):
            self.entries[field].delete(0, tk.END)
            self.entries[field].insert(0, item[i])

        self.subscription_var.set(item[len(self.entries)])
        self.employment_var.set(item[len(self.entries) + 1] == "Employed")

    def edit_data(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select an entry to edit.")
            return

        new_values = [self.entries[field].get() for field in self.entries]
        new_values.append(self.subscription_var.get())
        new_values.append("Employed" if self.employment_var.get() else "Unemployed")

        try:
            wb = openpyxl.load_workbook(self.filepath)
            sheet = wb.active
            row_index = self.tree.index(selected[0]) + 2
            for i, value in enumerate(new_values, start=1):
                sheet.cell(row=row_index, column=i, value=value)
            wb.save(self.filepath)
            wb.close()

            self.tree.item(selected[0], values=new_values)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update data: {e}")

    def delete_data(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select an entry to delete.")
            return

        try:
            wb = openpyxl.load_workbook(self.filepath)
            sheet = wb.active
            row_index = self.tree.index(selected[0]) + 2
            sheet.delete_rows(row_index)
            wb.save(self.filepath)
            wb.close()

            self.tree.delete(selected[0])
        except Exception as e:
            messagebox.showerror("Error", f"Failed to delete data: {e}")

    def format_headers(self, sheet):
        header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        bold_font = Font(bold=True)
        for col_num, header in enumerate(self.columns, start=1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = bold_font
            cell.fill = header_fill
            sheet.column_dimensions[cell.column_letter].width = len(header) + 5

if __name__ == "__main__":
    root = tb.Window(themename="darkly")
    app = ExcelApp(root)
    root.mainloop()
